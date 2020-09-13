from .forms import *
from .models import Document,ProcessDocument, Question, Subject, Topic, User

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Count
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect,get_object_or_404


from itertools import islice, chain
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json
import pandas as pd
import random

###################################
######## SUBJECT SECTION ##########
###################################
######### CONTAINS ALL ############
######### VIEWS RELATING ##########
########## TO SUBJECT ############
###################################

def index(request): #TO CHANGE TO PROPER NAME
    return render(request, 'questions/landingpage.html')

def subject(request):
    subjects = Subject.objects.all().order_by("subjectname")
    return render(request, 'questions/subject.html', {'subjects':subjects})

@login_required
@permission_required('questions.add_subject',raise_exception=True)
def addsubject(request):   
    if request.method == 'POST':       
        form = AddSubjectForm(request.POST)
        subject_name_entry = request.POST['subjectname']
        subject_name = request.POST['subjectname'].lower()        
        all_subjects =  [x.upper() for x in  Subject.objects.all().values_list('subjectname',flat=True)]
        
        if subject_name.upper() in all_subjects:
            messages.error(request, '<b>Error: </b><em>{}</em> already exists. Validation is case-insensitive (e.g. it and IT will be treated the same for validation against existing subjects).'.format(subject_name_entry))
        
        elif form.is_valid():
            upload_data = form.save(commit=False)
            upload_data = form.save()
            messages.success(request, '<b>Success</b>: </b><em>{}</em> has been added successfully and is now available for use.'.format(subject_name_entry))
    
        else:
            messages.error(request, '<b>Error:</b> The value entered "{}" is invalid. This may be due to an character or a processing error.<br>Please try again. If you encounter further errors please contact a staff member.'.format(subject_name_entry))
        
        return redirect('questions:addsubject')

    else:
        form = AddSubjectForm()
        return render(request, 'questions/subject_add.html', {'form': form, })

def editSubject(request, subjectname):
    subject = get_object_or_404(Subject,subjectname=subjectname)
    
    if request.method =="POST":
        form = EditSubjectForm(request.POST,instance=subject)
        new_subjectname = request.POST['subjectname']
        all_subjects =  [x.upper() for x in  Subject.objects.all().values_list('subjectname',flat=True)]
        
        if new_subjectname.upper() in all_subjects:
            messages.error(request, '<b>Error: </b><em>{}</em> already exists. Validation is case-insensitive (e.g. it and IT will be treated the same for validation against existing subjects).'.format(new_subjectname))
            new_subjectname = subjectname
        
        elif form.is_valid():
            form.save()
            messages.success(request, '<b>Success</b>: </b><em>{}</em> has been updated successfully'.format(new_subjectname))

        return redirect('questions:editsubject', subjectname=new_subjectname)
    
    else:
        form = EditSubjectForm(instance = subject)

    return render(request, 'questions/subject_edit.html', {'subject':subject,'subjectname':subjectname, 'form':form})


@permission_required('questions.delete_subject',raise_exception=True)
def deleteSubject(request, subjectname):
    
    from .utils import createTicketForSubjectDelete
    subject = Subject.objects.get(subjectname=subjectname)
    raiseTicket = createTicketForSubjectDelete(subject,request.user)
    if raiseTicket==201:
        messages.success(request, "<b>Success:</b> The subject '<em>{}</em>' has been flagged for deletion.".format(subject))
    else:
        messages.error(request, "<b>Error:</b> The subject '<em>{}</em>' could not be flagged for deletion - you may have already requested its deletion. Please try again in 60 seconds, or contact a Staff Member.".format(subject))
   
    return redirect("questions:editsubject",subjectname=subjectname)

###################################
######## TOPIC SECTION ##########
###################################
######### CONTAINS ALL ############
######### VIEWS RELATING ##########
########## TO TOPIC ############
###################################
    
def topic(request):    
    topics = Topic.objects.all()
    return render(request, 'questions/topic.html', {'topics':topics})    

@login_required
@permission_required('questions.add_topic',raise_exception=True)
def addtopic(request):
     
    if request.method == 'POST':       
        form = AddTopicForm(request.POST)
        topic_name = request.POST['topicname']
        all_topics =  [x.upper() for x in  Topic.objects.all().values_list('topicname',flat=True)]
        
        if topic_name.upper() in all_topics:
            messages.error(request, '<b>Error: </b><em>{}</em> already exists. Validation is case-insensitive (e.g. C1 and c1 will be treated the same for validation against existing topics).'.format(topic_name))                  
        
        elif form.is_valid():
            upload_data = form.save(commit=False)
            upload_data = form.save()
            messages.success(request, '<b>Success</b>: </b><em>{}</em> has been added successfully and is now available for use.'.format(topic_name))
            
        else:
            messages.error(request, '<b>Error:</b> The value entered "{}" is invalid. This may be due to an character or a processing error.<br>Please try again. If you encounter further errors please contact a staff member.'.format(topic_name))

        return redirect('questions:addtopic')
    
    else:
        form = AddTopicForm()
        return render(request, 'questions/topic_add.html', {'form': form, })

@login_required
@permission_required('questions.change_topic',raise_exception=True)
def editTopic(request, topicid):
    topic = get_object_or_404(Topic,pk=topicid)
    
    if request.method =="POST":
        form = EditTopicForm(request.POST,instance=topic)
        new_topicname = request.POST['topicname']
        all_topics =  [x.upper() for x in  Topic.objects.all().values_list('topicname',flat=True)]
        
        if new_topicname.upper() in all_topics:
            messages.error(request, '<b>Error: </b><em>{}</em> already exists. Validation is case-insensitive (e.g. C1 and c1 will be treated the same for validation against existing topics).'.format(new_topicname))
        
        elif form.is_valid():
            form.save()
            messages.success(request, '<b>Success</b>: </b><em>{}</em> has been updated successfully.'.format(new_topicname))
            
        return redirect('questions:edittopic', topicid=topicid)
    
    else:
        form = EditTopicForm(instance = topic)

    return render(request, 'questions/topic_edit.html', {'topic':topic,'topicid':topicid, 'form':form})


@permission_required('questions.delete_topic',raise_exception=True)
def deleteTopic(request, topicid):
    from .utils import createTicketForTopicDelete
    topic = Topic.objects.get(pk=topicid)
    raiseTicket = createTicketForTopicDelete(topic,request.user)
    if raiseTicket==201:
        messages.success(request, "<b>Success:</b> The topic '<em>{}</em>' has been flagged for deletion.".format(topic))
    else:
        messages.error(request, "<b>Error:</b> The topic '<em>{}</em>' could not be flagged for deletion - you may have already requested its deletion. Please try again in 60 seconds, or contact a Staff Member.".format(topic))

    return redirect("questions:edittopic",topicid=topicid)


def topicdetail(request, topicid):
    try:
        
        #questions_linked_to_topic = Question.objects.filter(topicid=topicid)
        topic  = Topic.objects.get(pk=topicid)
        questions_linked_to_topic = topic.question_set.all()
        subjects = set()
        for s in questions_linked_to_topic:
            subjects.add((int(s.subjectid_id), s.subjectid))
        subject = sorted(subjects)
    except Topic.DoesNotExist:
        raise Http404("The Topic does not exist or is not currently accessible to you.")
    return render(request, 'questions/topicdetail.html', {'topic':topic, 'question': questions_linked_to_topic, 'subject' : subject})
    #return HttpResponse("Looking at topic details for TopicID: {} ".format(topicid))
    
###################################
######## QUESTIONS SECTION ########
###################################
######### CONTAINS ALL ############
######### VIEWS RELATING ##########
########## TO QUESTIONS ###########
###################################    
    
def question(request):
    question_count = Question.objects.values('subjectid', 'subjectid__subjectname').order_by('subjectid').annotate(dcount=Count('subjectid'))
    return render(request, 'questions/questions.html', {'question_count':question_count})

@login_required
@permission_required('question.view_question',raise_exception=True)
def questionlist(request,subjectname):
    try:
        question  = Question.objects.filter(subjectid__subjectname=subjectname)        
    except Question.DoesNotExist:
        raise Http404("The Question does not exist or it not currently accessible to you.")

    if len(question) == 0:
        raise Http404("The Question does not exist or it not currently accessible to you.")

    return render(request, 'questions/questionsubject.html', {'question':question} )
 
###################################
######## UPLOAD SECTION ########
###################################
######### CONTAINS ALL ############
######### VIEWS RELATING ##########
########## TO UPLOAD ###########
###################################    
@login_required
@permission_required('questions.add_document',raise_exception=True)
def fileupload(request):
    uploader = User.objects.get(pk=request.user.id)
    if request.method == 'POST':       
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            upload_data = form.save(commit=False)
            upload_data.uploaded_by = uploader
            upload_data = form.save()
            #form.user = request.user
            
            request.session['filepath'] = str(upload_data)
            # sendMessage(request,getSystemID(),request.user.id,"File Upload",fileUploadMessage(request,request.session['filepath']))
            return redirect('questions:fileuploadsuccess')
    else:
        form = DocumentForm()
    return render(request, 'questions/model_form_upload.html', {'form': form, })



@login_required
@permission_required('questions.add_document',raise_exception=True)
def fileupload_success(request):

    if "filepath" not in request.session:
        return redirect('questions:fileupload')
    else:
        #return HttpResponse(settings.MEDIA_URL)
        uploaded_url = request.session['filepath']
        media_url = str(settings.MEDIA_URL[1:])
        return render(request, 'questions/fileupload_success.html', {'full_url': media_url + uploaded_url})

@login_required
@permission_required('upload.delete_document', raise_exception=True)
def fileDelete(request,fileid):
    
    file = Document.objects.get(pk=fileid)

    if request.user.is_superuser or request.user==file.uploaded_by:
        messages.success(request, "<b>Success:</b> The file '<em>{}</em>' has been deleted.".format(file.document))
        # sendMessage(request,getSystemID(),request.user.id,"File Deletion ({})".format(file.document),fileDeleted(request,file.document))

        file.delete()

   
    return redirect('questions:fileuploadprocesslist')
    
###################################
######## PROCESS SECTION ##########
###################################
######### CONTAINS ALL ############
######### VIEWS RELATING ##########
########## TO PROCESS #############
###################################
@login_required
@permission_required('questions.add_processdocument',raise_exception=True)
def fileupload_process(request, document_id):

    document = Document.objects.get(pk=document_id)
    file_details = documentExists(document_id)
    subject = document.subject
    str_document_name = str(document)
    document_name = str_document_name[str_document_name.rfind("/")+1:] 
      
    if file_details['exists'] == False:
        return render(request,'questions/fileupload_process_document.html',{'document':document})
        #return HttpResponse("<h3>Unable to Process</h3>" + (document.get_file_name()) + "<br> The file is not available - it may have been moved or deleted.<br>")
    else:
        response = HttpResponse()
    
        ####################################################
        ############## Validation Check ####################
        ############## If this fails, then set the document 
        ############# to be processed by System ############
        ####################################################

        ##### Validation - Check if the document has already been processed #####
        
        if ProcessDocument.objects.filter(document_id=document).exists():
            messages.error(request, '<b>Error:</b> {} appears to have already been processed. Please check the table to see if processing was successful.'.format(document_name))
            return redirect('questions:fileuploadprocesslist')

        ##### Check if the document can be opened #####
        try:
            data = readExcel(file_details['path'])
        except InvalidFileException as e:                      
            prevent_reprocessing = ProcessDocument.objects.create(document_id=document,processed_by=get_object_or_404(User,pk=42))
            prevent_reprocessing.save()
            
            messages.error(request, '<b>Error:</b> {} is in an unsupported format. Supported formats are: .xlsx,.xlsm,.xltx,.xltm,.csv,.txt.\nPlease try again.'.format(document_name))
            # sendMessage(request,getSystemID(),request.user.id,"File Processing Error ({})".format(document_name),fileProcessError(request,document))
            return redirect('questions:fileuploadprocesslist')
        
        
        #####################################
        ####### BEGIN TOPIC CHECKS #######
        #####################################
        ## Iterate through the topics to see which ones exist, and add it to a new list so we can create it.
        existing_topics = list(Topic.objects.filter(topicname__in=data['checkTopic']).values_list('topicname', flat=True))
        topics_to_be_created = [topic for topic in data['checkTopic'] if topic not in existing_topics]

        ## If topics_to_be_created is not empty (ie len >0) then we need to add things to the database, otherwise we can proceed
        
        if len(topics_to_be_created)>0:        
            add_topic = [Topic(topicname=topics_to_be_created[i]) for i in range(len(topics_to_be_created))]
            try:
               create_topics = Topic.objects.bulk_create(add_topic)              
            except Exception as e:  # TODO add better error handling - probably want to return a message
                raise Exception ("An error has occurred. No topics were created. \n{}".format(e))


        #####################################
        ####### BEGIN QUESTION CHECKS #######
        #####################################
        create_questions = []
        debug1 = []
        debug2 = []
        debug3 = []

        ## Query the question table to see if there are any questions that already exist with the topics for the uploading subject- this will be for existing topics.
        ## We can access the information using data['checkTopic']
        ## ruestion_topics = list(Question.objects.filter(topicid__topicname__in = question_dict.keys(),subjectid=subject).values_list('topicid__topicname',flat=True))
        check_topic_has_questions = set(Question.objects.filter(topicid__topicname__in = data['checkTopic'],subjectid=subject).values_list('topicid__topicname',flat=True))
       
        ##### CASE: ALL TOPICS HAS NO QUESTIONS IN DATABASE #####
        ## If check_topic_has_questions is empty that means that no topics in the file  have any questions associated to it in the DB, therefore it's probably logical to add all the questions with NO further validation checks.
        if len(check_topic_has_questions) == 0:
            for topic in data['questions']:
                topic_id = getTopicID(topic)
                for question in data['questions'][topic]:
                    create_questions.append(Question(questiontext=question[0],questionanswer=question[1], subjectid=subject, topicid=Topic(topicid=topic_id)))
                    ##@DEBUG
                    debug1.append([question[0],question[1], subject, topic_id])
       
        ##### CASE: TOPIC HAS QUESTIONS IN DATABASE #####
        ## a) if check_toopic_has_questions is not empty, but there are no associated questions for a topic, then we can add the questions directly to the database, regardless whether it exists for another topc - bulk_create will add all or none rows, so it makes sense to split this up rather than catching an exception later.    
        elif len(check_topic_has_questions) >0:
             #Do a set comparison i.e. set(a) - set(b)
            topics_with_no_questions = set(data['checkTopic'])  - check_topic_has_questions
            for topic in topics_with_no_questions:
                topic_id = getTopicID(topic)
                for question in data['questions'][topic]:
                    create_questions.append(Question(questiontext=question[0],questionanswer=question[1], subjectid=subject, topicid=Topic(topicid=topic_id)))
                    ##@@DEBUG
                    debug2.append([question[0],question[1], subject, topic_id])

        ## b)if check_topic_has_questions is not empty, then we need to check if the question+topic+subject combination is unique. Topic is the key of the data['questions'] dictionary
        for topic in check_topic_has_questions:
            topic_id = getTopicID(topic)

            # Once we have the topicID, we need to check if any of the questions associated to the topic already exist in the database. Do this by retrieving all questions
            # for that topic + subject combination 
            check_questions_for_topic = list(Question.objects.filter(topicid__topicname__in = data['checkTopic'],subjectid=subject).values_list('topicid','questiontext','questionanswer'))
            
            # Now check each question (accessing it through each topic key) from the spreadsheet against check_questions_for_topic
            # If it exists in both, then add it to a new list that we can access.
            # Then grab all the values from each question in the file and add it to another list in a specific format.
            # Then do a diff to find the unique questions to be added
            for question in data['questions'][topic]:
                question_exists = [[item[1],item[2]] for item in check_questions_for_topic]
                first_tuple_list = []
                
                for k,v in data['questions'].items():
                    # if the value is a list (which we expect it to be then iterate through and create a list to convert to tuples later)
                    first_tuple_list = [tuple(item) for item in v if isinstance(v, list)]
                        
                second_tuple_list = [tuple(lst) for lst in question_exists]

                ## Convert the first and second lists to sets and then diff them from each other to get unique questions.                    
                first_set = set(map(tuple, first_tuple_list))
                second_set = set(map(tuple, second_tuple_list))
                unique_questions_for_topic = list(set(first_set) - set(second_set))
            
            ###############################
            ### Now put it all together ###
            ###############################
            for uq in unique_questions_for_topic:
                ##@DEBUG
                debug3.append([uq[0],uq[1], subject, topic_id])
                create_questions.append(Question(questiontext=uq[0],questionanswer=uq[1], subjectid=subject, topicid=Topic(topicid=topic_id)))
        
        #### BEGIN FORM SECTION AND PROCESS####        
        if request.method == 'POST':
           
            form = ProcessDocumentForm(request.POST)
            if form.is_valid():
                Question.objects.bulk_create(create_questions, batch_size=None, ignore_conflicts=True)
                processor_user = User.objects.get(pk=request.user.id)
                document_details = Document.objects.get(pk=document_id)
                ##### Process the actual questions into databasse
                process_data = form.save(commit=False)
                process_data.processed_by = processor_user
                process_data.document_id = document_details
                
                process_data = form.save()
                messages.info(request, '<b>Success:</b> The file has been processed and the questions have been added to the selected subject.')
                # sendMessage(request,getSystemID(),request.user.id,"Processed File",fileProcessMessage(request,document))

                return redirect('questions:fileuploadprocesslist')
        else:
            form = ProcessDocumentForm()

            context = {
            'document_id' : document_id,
            'file_name' : document.get_file_name(),
            'data' : data['df'].head().to_html(index=False),    #index=False
            'subject' : subject,
            'form' : form
            }

            return render(request,'questions/fileupload_process_document.html',context)

##############################
###### HELPER FUNCTIONS ######
##############################

def documentExists(document_id):
    document = Document.objects.get(pk=document_id)
    full_path = (settings.MEDIA_ROOT + document.get_file_path() + document.get_file_name())
    document_exists = os.path.exists(full_path)    
    return {'exists': document_exists,'path': full_path}

def getTopicID(topic_name):
    topic_id = Topic.objects.get(topicname=topic_name).topicid
    return topic_id        

def readExcel(file_path):
    wb = load_workbook(file_path)
    ws = wb['Questions']
    
    #DataFrame Stuff
    data  = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols).head()
    
    #Die if row count is low (assuming title row) - turn this into a try
    if ws.max_row<2:
        raise Exception("Too short")
    
    question_dict = {}
    check_topic_exists = set()

    for count, row in enumerate(ws.iter_rows(min_row=2, min_col=1, values_only=True)):
        output = count
        question_text = row[1]
        answer_text = row[2]
        topic_text = row[3]
        
        
        if topic_text is not None:
            check_topic_exists.add(topic_text)
        
            if topic_text not in question_dict:
                question_dict[topic_text] =[ [row[1],row[2]]]
            else:
                question_dict[topic_text].extend([[row[1],row[2]]])
    
    return {'df':df, 'questions': question_dict, 'checkTopic': check_topic_exists}

@login_required
@permission_required('questions.view_processdocument',raise_exception=True)
def file_processlist(request):  
    if request.user.is_superuser:
        files = Document.objects.all()
    else:  
        files = Document.objects.filter(uploaded_by=request.user)

    processed_document = ProcessDocument.objects.filter(document_id__in=Document.objects.all()).values_list('document_id', flat=True)
    invalid_document = ProcessDocument.objects.filter(document_id__in=Document.objects.all(),processed_by=42).values_list('document_id', flat=True)




    return render(request, 'questions/process.html', {'files':files, 'docs': processed_document, 'invalid':invalid_document})

     
    
##########################################
##########################################
##########################################
####################################################
######### SELECT QUESTIONS FOR RANDOMNESS ##########
####################################################
# @login_required
def makeQuestion(request):
        if request.method == 'POST':
            
            form = QuestionForm(request.POST)
 
            response_data = {}
            response_data['topicid'] = request.POST.getlist('topicid[]')
            response_data['subjectid'] = request.POST['subjectid']
            response_data['choice_field'] = request.POST['customRadio']
            # response_data['choice_field'] = request.POST['choice_field']
            if 'number_of_questions' in request.POST:
                response_data['number_of_questions'] = request.POST['number_of_questions']
            
            request.session['questiondata'] = response_data
            
            return redirect('questions:createquestions')
        else:
            form = QuestionForm()

        # return render(request, 'upload/question_form.html', {'form': form, })
        return render(request, 'questions/question_generator_form.html', {'form': form, })

# @login_required
def createQuestions(request):


    if "questiondata" not in request.session:
        return redirect('questions:subject_add')
    else:
        subject_id = request.session['questiondata']['subjectid']
        topic_id = request.session['questiondata']['topicid']
        choice_field = request.session['questiondata']['choice_field']
    
        subject_name = Subject.objects.get(pk=subject_id) 
        #choice 1 'oneFromEach'  - 1 Question From Each Selected Topic 
        #choice 2 - 'tenFromEach'  ALl Questions For subject by topic   
        #choice 3  - 'fromMany' ANy Questions from selected topics
        
        # raise Exception()
        if choice_field == 'fromMany':
            number_of_questions = str(request.session['questiondata']['number_of_questions'])
        elif choice_field=='tenFromEach':
            request.session['subject'] = subject_id
            subject =  request.session['subject']
            return render(request,'questions/createquestions.html', {'subject':subject, 'choice_field' : choice_field, 'subject_name':subject_name })
        else:
            number_of_questions = str(1)
          
        

        question_list = []
        
        for topic in topic_id:

            question_dict = {}
            question_dict[topic] = []
            question = Question.objects.filter(subjectid = subject_id, topicid = topic)

            
            for q in question:
                # raise Exception(topic, q.questiontext, q.questionanswer)
                question_dict[topic].append([str(q.topicid),q.questiontext,q.questionanswer])
            question_list.append(question_dict)
        del(question_dict)

        

        question_output = {}
        question_multi_output = []

   

        for topics in question_list:
            for k,v in topics.items():
                question_output[k] = []
                if number_of_questions=='1' and choice_field=='oneFromEach':
                    question_output[k].append(random.choice(v))
                    qm = None
                    qm_json = question_output   
                else:
                    for topic in v:
                        if topic not in question_multi_output:
                            question_multi_output.append(topic)
        
        
        if question_multi_output:
            #randomise the list 
            random.shuffle(question_multi_output)
            qm = []

            ## Do a check to make sure that if we have fewer questions in the bank then specified, 
            ## return them all

            if len(question_multi_output) < (int(number_of_questions)):
                for x in range(0,len(question_multi_output)):
                    qm.append(question_multi_output[x])
                    question_output = None
                messages.warning(request, '<b>Warning:</b> The number of questions requested was greater than the total number of questions available for the selected subject and topic(s). Therefore, the  system has returned all possible questions. ')
                
            else:
                for x in range(0, int(number_of_questions)):
                    qm.append(question_multi_output[x])
                    question_output = None  

            qm_json = json.dumps(qm)   
        
        subject_name = Subject.objects.get(pk=subject_id)       

        context = {
            'subject' : subject_id,
            'subject_id' : subject_id,
            'subject_name': subject_name,
            'topic_id' : topic_id, 
            'choice_field' : choice_field, 
            'number_of_questions' : int(number_of_questions),
            'qo':question_output,
            'qm':qm,
            'qm_json': qm_json,
        }

        return render(request, 'questions/createquestions.html', context)

def load_topics(request):
    subject_id = request.GET.get('subject')
    topics = Question.objects.filter(subjectid=subject_id).order_by('topicid__topicname').values_list('topicid_id','topicid__topicname', flat=False).distinct()
    question_count =  Question.objects.filter(subjectid=subject_id).count()
    
    top_topics = Question.objects.filter(subjectid=subject_id).values_list("topicid__topicname").annotate(total=Count("topicid__topicname")).order_by('-total')
    top_ten_topic = "{}".format(",".join("'" + t[0] + "'" for t in top_topics))
    top_ten_totals = [topic[1] for topic in top_topics]

    context = {
        'top_topics' : str(top_ten_topic),
        'totals' : top_ten_totals,
        'topics':topics,
        'question_count': question_count
    }

    return render(request, 'questions/topic_dropdown_list_options.html', context)
